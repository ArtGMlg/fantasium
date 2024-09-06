import os
from fastapi import FastAPI, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from starlette.responses import RedirectResponse
from dotenv import load_dotenv
from fantasium.text_to_speech import TtsModel
from fantasium.chat_tokener import Tokener
from fantasium.schemas import Message
from fantasium.illustrator import Text2ImageAPI
from fantasium.transcriber import Transcriber
from fantasium.writer import GigachatInference
from fantasium.filter import Filter
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from settings import Settings

app = FastAPI()
settings = Settings()
tokener = Tokener(settings.Gigachat_KEY)
ttsmodel = TtsModel()
chatmodel = GigachatInference(settings.Story_len)
transcribe = Transcriber()
filter = Filter()
image_api = Text2ImageAPI('https://api-key.fusionbrain.ai/', settings.Kandinsky_API_KEY, settings.Kandinsky_SECRET_KEY)
model_id = image_api.get_model()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
    allow_credentials=True
)


# Redirect / -> Swagger-UI documentation
@app.get("/")
def main_function():
    """
    # Redirect
    to documentation (`/docs/`).
    """
    return RedirectResponse(url="/docs/")

@app.post("/generate_complete")
def generate_complete(message: Message):
    message = filter.filter_request(message)
    token = tokener.get_token()
    story_content, end_of_story = chatmodel.get_text(message.message, message.chat_id, token)
    audio_data = ttsmodel.get_audio(story_content)

    return {
        "story": story_content,
        "end_of_story": end_of_story,
        "transcription": audio_data,
    }

@app.post("/illustrator")
def send_illustator(model: Message):
    uuid = image_api.generate(filter.filter_request(model.message), model_id)
    images = image_api.check_generation(uuid)

    return {"image": images[0]}


@app.post('/transcribe')
async def upload_file(file: UploadFile = File()):
    context = await file.read()

    with open('temp_file.wav', 'wb') as temp_file:
        temp_file.write(context)
        result = transcribe.process('temp_file.wav')

    return {"transcription": result}


@app.post("/upload/")
async def upload_file(file: UploadFile = File(...)):
    # Сохранение загруженного файла
    file_location = f"temp_{file.filename}"
    with open(file_location, "wb") as f:
        f.write(await file.read())
    
    # Открываем существующий файл Excel
    wb = load_workbook(file_location)
    ws = wb.active  # Работаем с активным листом

    # Определяем диапазоны данных (предположим, данные в колонках A, B и C)
    data_ref = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=ws.max_row)
    categories_ref = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

    # Создаем диаграмму
    chart = BarChart()
    chart.title = "Диаграмма на основе загруженных данных"
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(categories_ref)

    # Устанавливаем диаграмму на лист
    ws.add_chart(chart, "E5")

    # Сохраняем измененный файл с диаграммой
    output_file = f"chart_{file.filename}"
    wb.save(output_file)

    # Удаляем временный загруженный файл
    os.remove(file_location)

    # Возвращаем файл с диаграммой
    return FileResponse(output_file,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=output_file)